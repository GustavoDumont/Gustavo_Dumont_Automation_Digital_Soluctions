from __future__ import annotations
import re
from pathlib import Path
import customtkinter as ctk
from tkinter import filedialog,messagebox
import fitz, pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Alignment
STATUS={'ativo':'Ativo','demitido':'Demitido','afastado':'Afastado','bloqueado':'Bloqueado','desmobilizado':'Desmobilizado','rescindido':'Rescindido','rescisao':'Rescisão','licenca':'Licença'}
DATE_SALARY=re.compile(r'(?P<adm>\d{2}/\d{2}/\d{4})(?:\s+(?P<dem>\d{2}/\d{2}/\d{4}))?\s+(?P<sal>\d{1,3}(?:\.\d{3})*,\d{2})')
def lines(words,tol=1.3):
    out=[]
    for w in sorted(words,key=lambda x:(x[1],x[0])):
        if not out or abs(out[-1][0]-w[1])>tol:out.append([w[1],[w]])
        else:out[-1][1].append(w)
    return [x[1] for x in out]
def extract(path):
    rows=[]
    with fitz.open(path) as doc:
        for page_no,page in enumerate(doc,1):
            ls=lines(page.get_text('words'))
            for i,line in enumerate(ls):
                text=' '.join(w[4] for w in line)
                m=re.match(r'^(\d{6})',text)
                if not m:continue
                chapa=m.group(1); name=[];role=[];status=''
                for w in line:
                    t=w[4].strip();n=t.lower()
                    if t==chapa:continue
                    if n in STATUS:status=STATUS[n]
                    elif w[0]<240:name.append(t)
                    elif w[0]<430:role.append(t)
                extra={'Admissão':'','Demissão':'','Salário':'','Contrato':''}
                for nxt in ls[i+1:i+5]:
                    nt=' '.join(w[4] for w in nxt);dm=DATE_SALARY.search(nt)
                    if dm:
                        cm=re.search(r'(?:\d{2}-)?(\d{8})',nt);extra={'Admissão':dm.group('adm'),'Demissão':dm.group('dem') or '','Salário':dm.group('sal'),'Contrato':cm.group(1) if cm else ''};break
                rows.append({'ID/chapa':chapa,'Nome':' '.join(name).strip(),'Salário':extra['Salário'],'Função':' '.join(role).strip(),'Situação':status,'Contrato':extra['Contrato'],'Admissão':extra['Admissão'],'Demissão':extra['Demissão'],'Página':page_no});break
    return pd.DataFrame(rows)
def save(df,path):
    df.to_excel(path,index=False,engine='openpyxl');wb=load_workbook(path);ws=wb.active;ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions
    for c in ws[1]:c.fill=PatternFill('solid',fgColor='1F5AA6');c.font=Font(color='FFFFFF',bold=True);c.alignment=Alignment(horizontal='center')
    widths=[14,34,16,28,16,14,14,14,10]
    for i,w in enumerate(widths,1):ws.column_dimensions[chr(64+i)].width=w
    wb.save(path)
class App(ctk.CTk):
    def __init__(self):
        super().__init__();self.title('Mapeamento FOPAG');self.geometry('720x420');self.pdf=ctk.StringVar();self._build()
    def _build(self):
        ctk.CTkLabel(self,text='Extrator de Efetivo da FOPAG',font=ctk.CTkFont(size=22,weight='bold')).pack(pady=22);ctk.CTkEntry(self,textvariable=self.pdf,width=580).pack(pady=8);ctk.CTkButton(self,text='Selecionar PDF',command=lambda:self.pdf.set(filedialog.askopenfilename(filetypes=[('PDF','*.pdf')]))).pack();ctk.CTkButton(self,text='Extrair e gerar Excel',fg_color='#F28C28',command=self.run).pack(pady=15)
    def run(self):
        try:
            df=extract(self.pdf.get())
            if df.empty:messagebox.showwarning('Sem dados','Nenhum registro encontrado.');return
            target=Path(self.pdf.get()).with_name(Path(self.pdf.get()).stem+'_efetivo.xlsx');save(df,target);messagebox.showinfo('Concluído',f'{len(df)} registros encontrados.')
        except Exception as e:messagebox.showerror('Erro',str(e))
if __name__=='__main__':App().mainloop()
